"""
1. import requests ==> 安裝request套件 (pip install requests), requests套件支援了HTTP的各種請求方法，其中GET與POST最常使用到
2. import re ==> 使用正規表達式 ( Regular expression )
3. import openpyxl ==> 安裝Excel操作套件 (pip install openpyxl)
4. import datetime ==> 偉了使用當下時間
"""
import os
import requests
import re
import openpyxl
import datetime

sitemap_source = 'https://www.esunbank.com/bank/sitemap.xml'


def export_urls():
    # 匯出的檔案絕對路徑
    excel_export_path = f'{get_mainpy_directory_path()}\\_testdata\\sitemap_export_Urls_list_{str(datetime.date.today())}.xlsx'

    # 撈網頁原始碼 > 存入字符串 "total_urls_list" 中
    sitemap_response = requests.get(sitemap_source)
    total_urls_list = sitemap_response.text

    # 使用正則表示式, 截取出字符串 "total_urls_list" 內所有URL格式字符, 存入字符串 "match_urls" 中
    url_result_List = []
    url_pattern = r'https?://\S+'
    match_urls = re.findall(url_pattern, total_urls_list)

    # 將字符串 "match_urls" 中每一筆字符做修改(移掉 </loc> 與 "), 添加進url_result_List 中
    for each_url in match_urls:
        url_result_List.append(each_url.replace("</loc>", '').replace('"', ''))

    # 將url_result_List 轉成集合set格式 (自動移除掉重複的字串, 集合不允许重复元素)
    unique_urls = set(url_result_List)

    # 引用excel套件
    excel_workbook = openpyxl.Workbook()
    # excel_workbook_structure_list 放入要分類的項目名稱
    excel_workbook_structure_list = ["/personal", "/personal/deposit", "/personal/loan", "/personal/credit-card",
                                     "/personal/wealth", "/personal/trust", "/personal/insurance", "/personal/lifefin",
                                     "/personal/apply", "/personal/event", "/small-business", "/corporate",
                                     "/digital", "/about", "/marketing", "/iframe/widget",
                                     "/error", "/bank-en", "/preview"]
    # 遍歷excel_workbook_structure_list中每一個項目
    for item in excel_workbook_structure_list:
        # excel工作簿中新增分頁與命名sheet名稱(ex: "/personal/wealth" ==> "personal_wealth")
        excel_worksheet = excel_workbook.create_sheet(title=item.replace("/", "_")[1:])
        index = 1
        for url in unique_urls:
            # 判斷當url內有包含item字串時, 將資料依序寫入該item名的分頁下
            if item in url:
                excel_worksheet.cell(row=index, column=1, value=url)
                index += 1
    # 移除第一個預設名為"sheet"的分頁
    excel_workbook.remove(excel_workbook.active)
    # 存檔
    excel_workbook.save(excel_export_path)


def get_mainpy_directory_path():
    # 获取当前脚本文件的绝对路径
    current_script_path = os.path.abspath(__file__)
    # 获取当前脚本文件所在目录的路径
    script_directory = os.path.dirname(current_script_path)
    # 获取 main.py 同一层目录的路径
    parent_directory = os.path.dirname(script_directory)
    return os.path.dirname(parent_directory)
